from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


def geral(lista):
    wb = Workbook()
    ws = wb.active
    a=3
    b=1
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width =30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width =30
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 30
    caminho='A1:E1'
    ws.merge_cells(caminho)
    ws['A1'].value = 'Lista Geral da corretora'
    ws.row_dimensions[1].height = 20
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=1, value="INDEX")
    ws.cell(row=2, column=2, value="CIDADE")
    ws.cell(row=2, column=3, value="OBRA")
    ws.cell(row=2, column=4, value="RESPONSAVEL")
    ws.cell(row=2, column=5, value="CNPJ")
    ws.cell(row=2, column=6, value="CLIENTE")
    ws.cell(row=2, column=7, value="VALOR")
    ws.cell(row=2, column=8, value="VALOR_TOTAL")
    ws.cell(row=2, column=9, value="CADASTRO")
    ws.cell(row=2, column=10, value="DATA")
    ws.cell(row=2, column=11, value="SERVIÇO")
    for i in lista:
        #print (i[0],i[1],i[2],i[3],i[4])
        ws.cell(row=a, column=b, value=i[0])
        ws.cell(row=a, column=b+1, value=i[1])
        ws.cell(row=a, column=b+2, value=i[2])
        ws.cell(row=a, column=b+3, value=i[3])
        ws.cell(row=a, column=b+4, value=i[4])
        ws.cell(row=a, column=b+5, value=i[5])
        ws.cell(row=a, column=b+6, value=i[6])
        ws.cell(row=a, column=b+7, value=i[7])
        ws.cell(row=a, column=b+8, value=i[8])
        ws.cell(row=a, column=b+9, value=i[9])
        ws.cell(row=a, column=b+10, value=i[10])
        
        a+=1
    ws.title = "RELACAO"
    wb.save("relacao.xlsx")
    return

